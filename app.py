# app.py - COMPLETE ALL-IN-ONE SOLUTION FOR STREAMLIT
import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime
import base64
from io import BytesIO
import warnings
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
import io
warnings.filterwarnings('ignore')

# ============================================
# SECTION 1: DATA PROCESSING CLASS
# ============================================
class ESTDataProcessor:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.exam_date = datetime(2022, 2, 4)
        self.load_all_sheets()
        
    def load_all_sheets(self):
        """Load all required sheets from the Excel file"""
        try:
            # Load sheets into DataFrames
            self.students = pd.read_excel(self.excel_path, sheet_name='Students')
            self.admissions = pd.read_excel(self.excel_path, sheet_name='Admissions')
            self.exam_notes = pd.read_excel(self.excel_path, sheet_name='Exam Notes')
            self.scores_est1 = pd.read_excel(self.excel_path, sheet_name='Scores EST I')
            self.scores_est2 = pd.read_excel(self.excel_path, sheet_name='Scores EST II')
            self.attendance_est1 = pd.read_excel(self.excel_path, sheet_name='Attendance EST I')
            self.attendance_est2 = pd.read_excel(self.excel_path, sheet_name='Attendance ESTII')
            self.est_scale = pd.read_excel(self.excel_path, sheet_name='EST Scale')
            self.note_types = pd.read_excel(self.excel_path, sheet_name='Note Types')
            
            # Clean column names
            for df in [self.students, self.admissions, self.exam_notes]:
                df.columns = df.columns.str.strip()
            
            # Clean email columns
            if 'email' in self.students.columns:
                self.students['email'] = self.students['email'].astype(str).str.strip().str.lower()
            if 'email' in self.admissions.columns:
                self.admissions['email'] = self.admissions['email'].astype(str).str.strip().str.lower()
            if 'email' in self.exam_notes.columns:
                self.exam_notes['email'] = self.exam_notes['email'].astype(str).str.strip().str.lower()
            
            return True
            
        except Exception as e:
            st.error(f"Error loading Excel file: {e}")
            return False
    
    def create_excel_with_formulas_bytes(self):
        """Create Excel file with formulas and return as bytes"""
        try:
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
            
            ws_final = wb.create_sheet('Final')
            
            # Headers
            headers = [
                'email', 'First Name', 'Last Name', 'gender', 'Date Of Birth',
                'Age on 4 Feb 2022', 'Registered EST I?', 'Registered EST II?',
                'Took EST I?', 'Took EST II?', 'Total # of Notes',
                'EST I Attendance Issues?', 'EST I Literacy', 'EST I Math',
                'EST I Total', 'EST II Math 1', 'EST II Biology', 'EST II Physics',
                'Grade', 'EST II Attendance Issues?', 'School Name', 'Cheating Flag'
            ]
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws_final.cell(row=1, column=col_idx, value=header)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Add data rows with formulas
            row_num = 2
            for idx, student in self.students.iterrows():
                email = str(student['email']).strip().lower() if pd.notna(student['email']) else ""
                
                # Basic info
                ws_final.cell(row=row_num, column=1, value=email)
                
                # Names
                name_parts = str(student['name']).split() if pd.notna(student['name']) else []
                first_name = name_parts[0] if name_parts else ''
                last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''
                ws_final.cell(row=row_num, column=2, value=first_name)
                ws_final.cell(row=row_num, column=3, value=last_name)
                
                # Gender
                gender = str(student['gender']).strip().title() if pd.notna(student['gender']) else ''
                ws_final.cell(row=row_num, column=4, value=gender)
                
                # Date of Birth
                dob = student['Date Of Birth']
                if pd.notna(dob):
                    ws_final.cell(row=row_num, column=5, value=dob)
                    # Age formula
                    age_formula = f'=DATEDIF(E{row_num}, DATE(2022,2,4), "Y") & "y " & DATEDIF(E{row_num}, DATE(2022,2,4), "YM") & "m " & DATEDIF(E{row_num}, DATE(2022,2,4), "MD") & "d"'
                    ws_final.cell(row=row_num, column=6, value=age_formula)
                
                # Registration formulas
                reg_est1_formula = f'=IF(COUNTIFS(Admissions!$B:$B, A{row_num}, Admissions!$C:$C, "EST I")>0, "Yes", "No")'
                ws_final.cell(row=row_num, column=7, value=reg_est1_formula)
                
                reg_est2_formula = f'=IF(COUNTIFS(Admissions!$B:$B, A{row_num}, Admissions!$C:$C, "EST II")>0, "Yes", "No")'
                ws_final.cell(row=row_num, column=8, value=reg_est2_formula)
                
                # Attendance formulas
                took_est1_formula = f'=IF(COUNTIFS(Attendance_EST_I!$A:$A, A{row_num}, Attendance_EST_I!$B:$B, "<>")>0, "Yes", "No")'
                ws_final.cell(row=row_num, column=9, value=took_est1_formula)
                
                took_est2_formula = f'=IF(COUNTIF(Attendance_ESTII!$A:$A, A{row_num})>0, "Yes", "No")'
                ws_final.cell(row=row_num, column=10, value=took_est2_formula)
                
                # Notes count
                notes_formula = f'=COUNTIF(Exam_Notes!$A:$A, A{row_num})'
                ws_final.cell(row=row_num, column=11, value=notes_formula)
                
                # Grade
                grade = student['Grade'] if pd.notna(student['Grade']) else ''
                ws_final.cell(row=row_num, column=19, value=grade)
                
                # School name formula
                school_formula = f'=IFERROR(INDEX(Admissions!$A:$A, MATCH(A{row_num}, Admissions!$B:$B, 0)), "")'
                ws_final.cell(row=row_num, column=21, value=school_formula)
                
                # Cheating flag formula
                cheating_formula = f'=IF(COUNTIFS(Exam_Notes!$A:$A, A{row_num}, Exam_Notes!$D:$D, "Cheating")>0, "Yes", "")'
                ws_final.cell(row=row_num, column=22, value=cheating_formula)
                
                # Score placeholders
                ws_final.cell(row=row_num, column=13, value=f'=IFERROR(VLOOKUP(A{row_num}, Scores_EST_I!$A:$E, 2, FALSE), "")')
                ws_final.cell(row=row_num, column=14, value=f'=IFERROR(VLOOKUP(A{row_num}, Scores_EST_I!$A:$E, 4, FALSE), "")')
                ws_final.cell(row=row_num, column=15, value=f'=IF(AND(NOT(ISBLANK(M{row_num})), NOT(ISBLANK(N{row_num}))), M{row_num}+N{row_num}, "")')
                
                row_num += 1
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            return output.getvalue()
            
        except Exception as e:
            st.error(f"Error creating Excel: {e}")
            return None
    
    def process_data(self):
        """Process data using Python calculations and return DataFrame"""
        try:
            final_data = []
            
            for idx, student in self.students.iterrows():
                email = str(student['email']).strip().lower() if pd.notna(student['email']) else ""
                
                # Basic info
                name_parts = str(student['name']).split() if pd.notna(student['name']) else []
                first_name = name_parts[0] if name_parts else ''
                last_name = ' '.join(name_parts[1:]) if len(name_parts) > 1 else ''
                
                # Age calculation
                age = self.calculate_age(student['Date Of Birth'])
                
                # Registration status
                registered_est1, registered_est2 = self.get_registration_status(email)
                
                # Attendance status
                took_est1, took_est2 = self.get_attendance_status(email)
                
                # Notes count
                total_notes = self.count_notes(email)
                
                # Attendance issues
                attendance_issues = self.check_attendance_issues(email)
                
                # Calculate scores
                est1_literacy = self.calculate_est1_literacy(email)
                est1_math = self.calculate_est1_math(email)
                est1_total = None
                if est1_literacy is not None and est1_math is not None:
                    est1_total = est1_literacy + est1_math
                
                est2_math = self.calculate_est2_subject(email, 'Math 1')
                est2_biology = self.calculate_est2_subject(email, 'Biology')
                est2_physics = self.calculate_est2_subject(email, 'Physics')
                
                # School name
                school_name = self.get_school_name(email)
                
                # Cheating flag
                cheating_flag = self.check_cheating(email)
                
                # Grade
                grade = student['Grade'] if pd.notna(student['Grade']) else ''
                
                # Compile row
                row = {
                    'email': email,
                    'First Name': first_name,
                    'Last Name': last_name,
                    'gender': student['gender'] if pd.notna(student['gender']) else '',
                    'Date Of Birth': student['Date Of Birth'],
                    'Age on 4 Feb 2022': age,
                    'Registered EST I?': registered_est1,
                    'Registered EST II?': registered_est2,
                    'Took EST I?': took_est1,
                    'Took EST II?': took_est2,
                    'Total # of Notes': total_notes,
                    'EST I Attendance Issues?': attendance_issues,
                    'EST I Literacy': est1_literacy,
                    'EST I Math': est1_math,
                    'EST I Total': est1_total,
                    'EST II Math 1': est2_math,
                    'EST II Biology': est2_biology,
                    'EST II Physics': est2_physics,
                    'Grade': grade,
                    'EST II Attendance Issues?': '',
                    'School Name': school_name,
                    'Cheating Flag': cheating_flag
                }
                
                final_data.append(row)
            
            df = pd.DataFrame(final_data)
            return df
            
        except Exception as e:
            st.error(f"Error processing data: {e}")
            return None
    
    # Helper methods
    def calculate_age(self, dob):
        """Calculate age as of Feb 4, 2022"""
        if pd.isna(dob):
            return ''
        
        try:
            if isinstance(dob, str):
                dob = datetime.strptime(dob.split()[0], '%Y-%m-%d')
            
            exam_date = datetime(2022, 2, 4)
            
            years = exam_date.year - dob.year
            months = exam_date.month - dob.month
            days = exam_date.day - dob.day
            
            if days < 0:
                months -= 1
                days += 30
            
            if months < 0:
                years -= 1
                months += 12
            
            return f"{years}y {months}m {days}d"
        except:
            return ''
    
    def get_registration_status(self, email):
        """Get registration status for a student"""
        if not email:
            return 'No', 'No'
        
        student_admissions = self.admissions[self.admissions['email'] == email]
        registered_est1 = 'Yes' if not student_admissions.empty and 'EST I' in student_admissions['exam'].values else 'No'
        registered_est2 = 'Yes' if not student_admissions.empty and 'EST II' in student_admissions['exam'].values else 'No'
        return registered_est1, registered_est2
    
    def get_attendance_status(self, email):
        """Get attendance status for a student"""
        if not email:
            return 'No', 'No'
        
        # EST I attendance
        est1_att = self.attendance_est1[self.attendance_est1['Email'] == email]
        took_est1 = 'No'
        if not est1_att.empty:
            time_in = est1_att['Time In'].iloc[0]
            if pd.notna(time_in):
                took_est1 = 'Yes'
        
        # EST II attendance
        est2_att = self.attendance_est2[self.attendance_est2['Email'] == email]
        took_est2 = 'Yes' if not est2_att.empty else 'No'
        
        return took_est1, took_est2
    
    def count_notes(self, email):
        """Count notes for a student"""
        if not email:
            return 0
        return len(self.exam_notes[self.exam_notes['email'] == email])
    
    def check_attendance_issues(self, email):
        """Check for attendance issues"""
        if not email:
            return ''
        
        notes = self.exam_notes[self.exam_notes['email'] == email]
        est1_att = self.attendance_est1[self.attendance_est1['Email'] == email]
        
        issues = []
        
        # Check if marked absent but has attendance
        if 'Absent' in notes['Note'].values:
            if not est1_att.empty:
                time_in = est1_att['Time In'].iloc[0]
                if pd.notna(time_in):
                    issues.append('Marked absent but attended')
        
        # Check if has attendance but no scores
        if not est1_att.empty:
            time_in = est1_att['Time In'].iloc[0]
            if pd.notna(time_in):
                scores = self.scores_est1[self.scores_est1['Email'] == email]
                if scores.empty or scores.isna().all().all():
                    issues.append('Attended but no scores')
        
        return '; '.join(issues) if issues else ''
    
    def check_cheating(self, email):
        """Check if student has cheating note"""
        if not email:
            return ''
        notes = self.exam_notes[self.exam_notes['email'] == email]
        return 'Yes' if 'Cheating' in notes['Note'].values else ''
    
    def get_school_name(self, email):
        """Get school name for a student"""
        if not email:
            return ''
        schools = self.admissions[self.admissions['email'] == email]['School Name']
        return schools.iloc[0] if not schools.empty else ''
    
    def calculate_est1_literacy(self, email):
        """Calculate EST I Literacy score"""
        if not email:
            return None
        
        scores = self.scores_est1[self.scores_est1['Email'] == email]
        if scores.empty:
            return None
        
        # Check for misconduct
        notes = self.exam_notes[self.exam_notes['email'] == email]
        has_misconduct = 'Misconduct' in notes['Note'].values
        
        # Get raw scores
        literacy1_raw = scores['EST I Literacy 1'].iloc[0] if pd.notna(scores['EST I Literacy 1'].iloc[0]) else 0
        literacy2_raw = scores['EST I Literacy 2'].iloc[0] if pd.notna(scores['EST I Literacy 2'].iloc[0]) else 0
        
        # Apply raises if no misconduct
        if not has_misconduct:
            literacy1_raw += 4
            literacy2_raw += 3
        
        # Scale scores
        literacy1_scaled = self.scale_score(literacy1_raw, 'EST I Literacy 1')
        literacy2_scaled = self.scale_score(literacy2_raw, 'EST I Literacy 2')
        
        if literacy1_scaled is not None and literacy2_scaled is not None:
            return (literacy1_scaled + literacy2_scaled) * 10
        
        return None
    
    def calculate_est1_math(self, email):
        """Calculate EST I Math score"""
        if not email:
            return None
        
        scores = self.scores_est1[self.scores_est1['Email'] == email]
        if scores.empty:
            return None
        
        # Check for misconduct
        notes = self.exam_notes[self.exam_notes['email'] == email]
        has_misconduct = 'Misconduct' in notes['Note'].values
        
        # Get raw scores
        math_no_calc = scores['EST I Math without Calculator'].iloc[0] if pd.notna(scores['EST I Math without Calculator'].iloc[0]) else 0
        math_with_calc = scores['EST I Math with Calculator'].iloc[0] if pd.notna(scores['EST I Math with Calculator'].iloc[0]) else 0
        
        # Apply raises if no misconduct
        if not has_misconduct:
            math_no_calc += 5
            math_with_calc += 1
        
        total_raw = math_no_calc + math_with_calc
        
        # Scale score
        return self.scale_score(total_raw, 'EST I Math')
    
    def calculate_est2_subject(self, email, subject):
        """Calculate EST II subject score"""
        if not email:
            return None
        
        if subject == 'Math 1':
            col = 'EST II Math 1'
            scale_key = 'EST II Math 1'
        elif subject == 'Biology':
            col = 'EST II Biology'
            scale_key = 'EST II Biology'
        elif subject == 'Physics':
            col = 'EST II Physics'
            scale_key = 'EST II Physics'
        else:
            return None
        
        scores = self.scores_est2[self.scores_est2['Email'] == email]
        if scores.empty or col not in scores.columns:
            return None
        
        raw_score = scores[col].iloc[0]
        if pd.isna(raw_score):
            return None
        
        # Check for misconduct
        notes = self.exam_notes[self.exam_notes['email'] == email]
        has_misconduct = 'Misconduct' in notes['Note'].values
        
        # Apply raise based on average
        if not has_misconduct:
            subject_scores = self.scores_est2[col].dropna()
            if len(subject_scores) > 0:
                avg_score = subject_scores.mean()
                raise_amount = max(1, min(5, int(avg_score / 10)))
                raw_score += raise_amount
        
        # Scale score
        return self.scale_score(raw_score, scale_key)
    
    def scale_score(self, raw_score, subject):
        """Scale raw score using EST Scale"""
        if pd.isna(raw_score):
            return None
        
        # Map subjects to scale columns
        scale_map = {
            'EST I Math': (0, 1),
            'EST I Literacy 1': (3, 4),
            'EST I Literacy 2': (6, 7),
            'EST II Math 1': (9, 10),
            'EST II Biology': (12, 13),
            'EST II Physics': (15, 16)
        }
        
        for key, (raw_col, scale_col) in scale_map.items():
            if key in subject:
                scale_df = self.est_scale.iloc[:, [raw_col, scale_col]].dropna()
                scale_df.columns = ['Raw', 'Scaled']
                
                # Find closest raw score
                scale_df = scale_df.sort_values('Raw')
                match = scale_df[scale_df['Raw'] <= raw_score]
                
                if match.empty:
                    return scale_df['Scaled'].iloc[0]
                else:
                    return match.iloc[-1]['Scaled']
        
        return None

# ============================================
# SECTION 2: DASHBOARD CLASS
# ============================================
class StreamlitESTDashboard:
    def __init__(self):
        self.processor = None
        self.df = None
        self.initialize_session_state()
    
    def initialize_session_state(self):
        """Initialize session state variables"""
        if 'data_loaded' not in st.session_state:
            st.session_state.data_loaded = False
        if 'df' not in st.session_state:
            st.session_state.df = None
        if 'excel_bytes' not in st.session_state:
            st.session_state.excel_bytes = None
    
    def load_data(self):
        """Load and process data"""
        if not st.session_state.data_loaded:
            with st.spinner("📊 Loading and processing data..."):
                try:
                    self.processor = ESTDataProcessor("Employment Test - Dataset - TASK A.xlsx")
                    if self.processor:
                        # Process data for dashboard
                        st.session_state.df = self.processor.process_data()
                        
                        # Create Excel with formulas
                        excel_bytes = self.processor.create_excel_with_formulas_bytes()
                        if excel_bytes:
                            st.session_state.excel_bytes = excel_bytes
                        
                        st.session_state.data_loaded = True
                        st.success(f"✅ Successfully loaded {len(st.session_state.df)} student records")
                        
                except Exception as e:
                    st.error(f"❌ Error loading data: {e}")
                    st.info("Using sample data for demonstration")
                    self.create_sample_data()
        else:
            self.df = st.session_state.df
    
    def create_sample_data(self):
        """Create sample data for demo"""
        np.random.seed(42)
        n_students = 50
        
        self.df = pd.DataFrame({
            'email': [f'est{i:04d}@gmail.com' for i in range(1000, 1000+n_students)],
            'First Name': ['Student'] * n_students,
            'Last Name': [f'{i}' for i in range(1, n_students+1)],
            'gender': np.random.choice(['Male', 'Female'], n_students),
            'Age on 4 Feb 2022': [f'{np.random.randint(15, 19)}y {np.random.randint(0, 12)}m {np.random.randint(0, 30)}d' for _ in range(n_students)],
            'Registered EST I?': np.random.choice(['Yes', 'No'], n_students, p=[0.9, 0.1]),
            'Registered EST II?': np.random.choice(['Yes', 'No'], n_students, p=[0.7, 0.3]),
            'Took EST I?': np.random.choice(['Yes', 'No'], n_students, p=[0.85, 0.15]),
            'Took EST II?': np.random.choice(['Yes', 'No'], n_students, p=[0.6, 0.4]),
            'Total # of Notes': np.random.randint(0, 5, n_students),
            'EST I Attendance Issues?': [''] * n_students,
            'EST I Literacy': np.random.normal(600, 100, n_students).clip(400, 800),
            'EST I Math': np.random.normal(650, 120, n_students).clip(400, 800),
            'EST I Total': np.random.normal(1250, 200, n_students).clip(800, 1600),
            'Grade': np.random.choice(['Grade 10', 'Grade 11', 'Grade 12'], n_students),
            'School Name': np.random.choice(['EST School 1', 'EST School 2', 'EST School 3', 
                                            'EST School 4', 'EST School 5'], n_students),
            'Cheating Flag': np.random.choice(['Yes', ''], n_students, p=[0.05, 0.95])
        })
        self.df['EST I Total'] = self.df['EST I Literacy'] + self.df['EST I Math']
        st.session_state.df = self.df
        st.session_state.data_loaded = True
    
    def run_dashboard(self):
        """Run the complete dashboard"""
        # Sidebar
        with st.sidebar:
            st.title("📊 EST Dashboard")
            st.markdown("---")
            
            # Load data button
            if not st.session_state.data_loaded:
                if st.button("🚀 Load Data", type="primary", use_container_width=True):
                    self.load_data()
                    st.rerun()
            else:
                st.success("✅ Data loaded")
            
            st.markdown("---")
            
            # Filters
            st.subheader("🔍 Filters")
            
            # School filter
            schools = ['All Schools'] + sorted(self.df['School Name'].dropna().unique().tolist())
            selected_school = st.selectbox("Select School", schools)
            
            # Grade filter
            grades = ['All Grades'] + sorted(self.df['Grade'].dropna().unique().tolist())
            selected_grade = st.selectbox("Select Grade", grades)
            
            # Exam filter
            exam_options = ['All Exams', 'EST I Only', 'EST II Only', 'Both Exams']
            selected_exam = st.selectbox("Exam Type", exam_options)
            
            # Score range filter
            if not self.df['EST I Total'].isna().all():
                min_score = int(self.df['EST I Total'].min())
                max_score = int(self.df['EST I Total'].max())
                score_range = st.slider(
                    "EST I Score Range",
                    min_value=min_score,
                    max_value=max_score,
                    value=(min_score, max_score)
                )
            
            # Cheating cases filter
            show_cheating_only = st.checkbox("Show Cheating Cases Only")
            
            st.markdown("---")
            
            # Export section
            st.subheader("📤 Export Data")
            
            if st.session_state.excel_bytes:
                st.download_button(
                    label="📥 Download Excel (with formulas)",
                    data=st.session_state.excel_bytes,
                    file_name="EST_Final_With_Formulas.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            if self.df is not None:
                csv = self.df.to_csv(index=False).encode('utf-8')
                st.download_button(
                    label="📄 Download CSV (cleaned data)",
                    data=csv,
                    file_name="final_data_cleaned.csv",
                    mime="text/csv",
                    use_container_width=True
                )
        
        # Apply filters
        filtered_df = self.apply_filters(
            selected_school, 
            selected_grade, 
            selected_exam,
            score_range if 'score_range' in locals() else None,
            show_cheating_only
        )
        
        # Main content
        st.markdown('<h1 class="main-header">🎯 EST Assessment Dashboard</h1>', unsafe_allow_html=True)
        
        # Key Metrics
        self.display_metrics(filtered_df)
        
        # Tabs
        tab1, tab2, tab3, tab4, tab5 = st.tabs([
            "📈 Overview", 
            "🏫 Schools", 
            "📊 Performance", 
            "⚠️ Issues", 
            "👨‍🎓 Students"
        ])
        
        with tab1:
            self.overview_tab(filtered_df)
        
        with tab2:
            self.schools_tab(filtered_df)
        
        with tab3:
            self.performance_tab(filtered_df)
        
        with tab4:
            self.issues_tab(filtered_df)
        
        with tab5:
            self.students_tab(filtered_df)
    
    def apply_filters(self, school, grade, exam, score_range, cheating_only):
        """Apply filters to data"""
        filtered = self.df.copy()
        
        # School filter
        if school != 'All Schools':
            filtered = filtered[filtered['School Name'] == school]
        
        # Grade filter
        if grade != 'All Grades':
            filtered = filtered[filtered['Grade'] == grade]
        
        # Exam filters
        if exam == 'EST I Only':
            filtered = filtered[filtered['Took EST I?'] == 'Yes']
        elif exam == 'EST II Only':
            filtered = filtered[filtered['Took EST II?'] == 'Yes']
        elif exam == 'Both Exams':
            filtered = filtered[(filtered['Took EST I?'] == 'Yes') & (filtered['Took EST II?'] == 'Yes')]
        
        # Score range filter
        if score_range is not None:
            filtered = filtered[
                (filtered['EST I Total'] >= score_range[0]) & 
                (filtered['EST I Total'] <= score_range[1])
            ]
        
        # Cheating filter
        if cheating_only:
            filtered = filtered[filtered['Cheating Flag'] == 'Yes']
        
        return filtered
    
    def display_metrics(self, data):
        """Display key metrics"""
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            total_students = len(data)
            st.metric("Total Students", total_students)
        
        with col2:
            est1_takers = data[data['Took EST I?'] == 'Yes'].shape[0]
            st.metric("EST I Takers", est1_takers)
        
        with col3:
            est2_takers = data[data['Took EST II?'] == 'Yes'].shape[0]
            st.metric("EST II Takers", est2_takers)
        
        with col4:
            if not data['EST I Total'].isna().all():
                avg_score = data['EST I Total'].mean()
                st.metric("Avg EST I Score", f"{avg_score:.0f}")
            else:
                st.metric("Avg EST I Score", "N/A")
        
        with col5:
            cheating_cases = data[data['Cheating Flag'] == 'Yes'].shape[0]
            st.metric("Cheating Cases", cheating_cases)
        
        st.markdown("---")
    
    def overview_tab(self, data):
        """Overview tab content"""
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("📊 EST I Score Distribution")
            if not data['EST I Total'].isna().all():
                fig = px.histogram(
                    data, 
                    x='EST I Total',
                    nbins=20,
                    title="Distribution of EST I Scores",
                    labels={'EST I Total': 'Score'},
                    color_discrete_sequence=['#3B82F6']
                )
                fig.update_layout(height=400)
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No EST I score data available")
        
        with col2:
            st.subheader("📋 Registration vs Attendance")
            
            reg_att_data = pd.DataFrame({
                'Category': ['EST I Registered', 'EST I Attended', 'EST II Registered', 'EST II Attended'],
                'Count': [
                    data[data['Registered EST I?'] == 'Yes'].shape[0],
                    data[data['Took EST I?'] == 'Yes'].shape[0],
                    data[data['Registered EST II?'] == 'Yes'].shape[0],
                    data[data['Took EST II?'] == 'Yes'].shape[0]
                ]
            })
            
            fig = px.bar(
                reg_att_data, 
                x='Category', 
                y='Count',
                title="",
                color='Category',
                color_discrete_sequence=['#3B82F6', '#10B981', '#6366F1', '#8B5CF6']
            )
            fig.update_layout(height=400, xaxis_tickangle=-45)
            st.plotly_chart(fig, use_container_width=True)
        
        # Data preview
        st.subheader("📋 Data Preview")
        st.dataframe(data.head(10), use_container_width=True)
    
    def schools_tab(self, data):
        """Schools analysis tab"""
        if data['School Name'].nunique() > 1:
            st.subheader("🏫 School Performance Comparison")
            
            # Calculate school statistics
            school_stats = data.groupby('School Name').agg({
                'email': 'count',
                'EST I Total': 'mean',
                'EST I Literacy': 'mean',
                'EST I Math': 'mean',
                'Took EST I?': lambda x: (x == 'Yes').sum(),
                'Took EST II?': lambda x: (x == 'Yes').sum(),
                'Cheating Flag': lambda x: (x == 'Yes').sum()
            }).round(2)
            
            school_stats.columns = [
                'Total Students', 'Avg Total Score', 'Avg Literacy', 'Avg Math',
                'EST I Takers', 'EST II Takers', 'Cheating Cases'
            ]
            
            st.dataframe(school_stats, use_container_width=True)
            
        else:
            st.info(f"Showing data for {data['School Name'].iloc[0] if not data.empty else 'selected school'}")
    
    def performance_tab(self, data):
        """Performance analysis tab"""
        st.subheader("📊 Performance Analysis")
        
        col1, col2 = st.columns(2)
        
        with col1:
            if not data['EST I Literacy'].isna().all() and not data['EST I Math'].isna().all():
                st.subheader("📈 Literacy vs Math Scores")
                fig = px.scatter(
                    data,
                    x='EST I Literacy',
                    y='EST I Math',
                    color='Grade',
                    title="Correlation between Literacy and Math Scores",
                    trendline="ols"
                )
                st.plotly_chart(fig, use_container_width=True)
        
        with col2:
            if not data['EST I Total'].isna().all():
                st.subheader("📊 Score Distribution by Grade")
                fig = px.box(
                    data,
                    x='Grade',
                    y='EST I Total',
                    color='Grade',
                    title="EST I Score Distribution by Grade"
                )
                st.plotly_chart(fig, use_container_width=True)
        
        # Top performers
        if not data['EST I Total'].isna().all():
            st.subheader("🏆 Top 10 Performers")
            top_performers = data.nlargest(10, 'EST I Total')[['First Name', 'Last Name', 'Grade', 'School Name', 'EST I Total']]
            st.dataframe(top_performers, use_container_width=True)
    
    def issues_tab(self, data):
        """Issues and quality control tab"""
        st.subheader("⚠️ Data Quality Issues")
        
        total_students = len(data)
        
        # Calculate issues
        attendance_issues = data[data['EST I Attendance Issues?'] != ''].shape[0]
        missing_scores = data['EST I Total'].isna().sum()
        cheating_cases = data[data['Cheating Flag'] == 'Yes'].shape[0]
        
        # Display metrics
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.metric("Attendance Issues", attendance_issues)
        
        with col2:
            st.metric("Missing Scores", missing_scores)
        
        with col3:
            st.metric("Cheating Cases", cheating_cases)
        
        with col4:
            completeness = ((total_students - missing_scores) / total_students * 100) if total_students > 0 else 0
            st.metric("Data Completeness", f"{completeness:.1f}%")
        
        # Detailed issues
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("🚨 Students with Attendance Issues")
            issues_df = data[data['EST I Attendance Issues?'] != ''][['First Name', 'Last Name', 'email', 'EST I Attendance Issues?']]
            if not issues_df.empty:
                st.dataframe(issues_df, use_container_width=True)
            else:
                st.success("No attendance issues found!")
        
        with col2:
            st.subheader("🚫 Cheating Cases")
            cheating_df = data[data['Cheating Flag'] == 'Yes'][['First Name', 'Last Name', 'email', 'School Name', 'Grade']]
            if not cheating_df.empty:
                st.dataframe(cheating_df, use_container_width=True)
            else:
                st.success("No cheating cases found!")
    
    def students_tab(self, data):
        """Students details tab"""
        st.subheader("👨‍🎓 Student Details")
        
        # Search and filter
        col1, col2 = st.columns(2)
        with col1:
            search_term = st.text_input("🔍 Search by Name or Email")
        with col2:
            sort_by = st.selectbox(
                "Sort by",
                ['EST I Total (High to Low)', 'EST I Total (Low to High)', 'Name (A-Z)', 'Name (Z-A)']
            )
        
        # Filter data
        filtered_data = data.copy()
        if search_term:
            filtered_data = filtered_data[
                filtered_data['email'].str.contains(search_term, case=False, na=False) |
                filtered_data['First Name'].str.contains(search_term, case=False, na=False) |
                filtered_data['Last Name'].str.contains(search_term, case=False, na=False)
            ]
        
        # Sort data
        if sort_by == 'EST I Total (High to Low)':
            filtered_data = filtered_data.sort_values('EST I Total', ascending=False)
        elif sort_by == 'EST I Total (Low to High)':
            filtered_data = filtered_data.sort_values('EST I Total', ascending=True)
        elif sort_by == 'Name (A-Z)':
            filtered_data = filtered_data.sort_values(['Last Name', 'First Name'])
        elif sort_by == 'Name (Z-A)':
            filtered_data = filtered_data.sort_values(['Last Name', 'First Name'], ascending=False)
        
        # Display student table
        st.dataframe(
            filtered_data[[
                'First Name', 'Last Name', 'email', 'Grade', 'School Name',
                'EST I Total', 'EST I Literacy', 'EST I Math',
                'Registered EST I?', 'Took EST I?', 'Cheating Flag'
            ]],
            use_container_width=True
        )

# ============================================
# SECTION 3: MAIN STREAMLIT APP
# ============================================
def main():
    """Main Streamlit application"""
    # Set page config
    st.set_page_config(
        page_title="EST Assessment System",
        page_icon="🎓",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    # Custom CSS
    st.markdown("""
    <style>
        .main-header {
            font-size: 2.5rem;
            color: #1E3A8A;
            text-align: center;
            margin-bottom: 2rem;
            font-weight: 700;
        }
        .stButton > button {
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            border: none;
            padding: 0.5rem 1rem;
            border-radius: 5px;
            font-weight: 600;
        }
        .stButton > button:hover {
            background: linear-gradient(135deg, #5a67d8 0%, #6b46c1 100%);
            color: white;
        }
    </style>
    """, unsafe_allow_html=True)
    
    # Title
    st.markdown('<h1 class="main-header">🎓 EST Assessment System</h1>', unsafe_allow_html=True)
    
    # Initialize dashboard
    dashboard = StreamlitESTDashboard()
    
    # Check if data is loaded
    if not st.session_state.data_loaded:
        # Show welcome screen
        col1, col2, col3 = st.columns([1, 2, 1])
        with col2:
            st.markdown("""
            <div style="text-align: center; padding: 2rem; background: #f8f9fa; border-radius: 10px;">
                <h2>📊 Welcome to EST Assessment System</h2>
                <p>This system processes EST exam data and provides interactive analytics.</p>
                <p>Click the button below to load and process the data.</p>
            </div>
            """, unsafe_allow_html=True)
            
            if st.button("🚀 Load & Process Data", type="primary", use_container_width=True, use_container_width=True):
                with st.spinner("Processing data..."):
                    dashboard.load_data()
                st.rerun()
            
            st.markdown("---")
            
            # Features
            st.subheader("✨ Features")
            col1, col2, col3 = st.columns(3)
            with col1:
                st.markdown("""
                **📈 Data Processing**
                - Excel with formulas
                - Cleaned CSV export
                - Score calculations
                """)
            with col2:
                st.markdown("""
                **📊 Analytics**
                - Interactive dashboard
                - School comparisons
                - Performance metrics
                """)
            with col3:
                st.markdown("""
                **📋 Reports**
                - Data quality checks
                - Attendance issues
                - Cheating detection
                """)
    else:
        # Run dashboard with loaded data
        dashboard.df = st.session_state.df
        dashboard.run_dashboard()

# Run the app
if __name__ == "__main__":
    main()
